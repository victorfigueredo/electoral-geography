import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Replace 'your_file_path.csv' with the path to your CSV file
csv_file_path = 'data/2022/sp_voting_type_tse.csv'
# Replace 'your_output_file_path.xls' with the path to your output Excel file
xls_file_path = 'data/2022/sp_voting_type_tse.xls'

# Read the CSV file with ';' as the delimiter and create a DataFrame
df = pd.read_csv(csv_file_path, sep=';')

# Group the DataFrame by 'voting_type'
grouped_df = df.groupby('voting_type')

# Create a new Excel Workbook
wb = Workbook()

# Loop through each group in the grouped DataFrame
for group_name, group_data in grouped_df:
    # Create a new sheet in the workbook with the group name
    ws = wb.create_sheet(title=group_name)

    # Write the group data to the sheet
    for row in dataframe_to_rows(group_data, index=False, header=True):
        ws.append(row)

# Remove the default empty sheet in the workbook
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Save the Excel file
wb.save(xls_file_path)

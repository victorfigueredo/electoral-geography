import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List
import os
from pathlib import Path

class ExportData:
    """
    Class to export Pandas DataFrame data to CSV and Excel files.
    """
    
    def __init__(self, dataframe: pd.DataFrame):
        """
        Args:
            dataframe (pd.DataFrame): The input dataframe that needs to be exported.
        """
        self.dataframe = dataframe

    def to_csv(self, file_path: str, delimiter: str = ',') -> None:
        """
        Export dataframe to CSV file.

        Args:
            file_path (str): The path where the CSV file will be saved.
            delimiter (str, optional): Character used to separate values. Default is ','.
        """
        # Create directory if it does not exist
        directory = Path(file_path).parent
        directory.mkdir(parents=True, exist_ok=True)
        
        # Now write the CSV
        self.dataframe.to_csv(file_path, sep=delimiter, index=False)


    def to_excel(self, file_path: str, sheet_name: str = 'Sheet1', group_by: List[str] = None) -> None:
        """
        Export dataframe to Excel file. Can export different groups to different sheets.

        Args:
            file_path (str): The path where the Excel file will be saved.
            sheet_name (str, optional): The name of the sheet where the data will be saved. Default is 'Sheet1'.
            group_by (List[str], optional): Column(s) to group by. Each group will be written to a different sheet. Default is None.
        """
        wb = Workbook()
        if group_by:
            grouped_df = self.dataframe.groupby(group_by)
            for group_name, group_data in grouped_df:
                ws = wb.create_sheet(title=str(group_name))
                for row in dataframe_to_rows(group_data, index=False, header=True):
                    ws.append(row)
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        else:
            ws = wb.create_sheet(title=sheet_name)
            for row in dataframe_to_rows(self.dataframe, index=False, header=True):
                ws.append(row)
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb['Sheet'])

        wb.save(file_path)
